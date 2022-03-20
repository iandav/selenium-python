import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl

# Set up openpyxl
path = 'C:/Users/Majestic/Desktop/data.xlsx'
workbook = openpyxl.load_workbook(path)
print(workbook.active)
sheet = workbook.active

# Read data
totalRows = sheet.max_row
totalColumns = sheet.max_column
print(totalRows)
print(totalColumns)

for r in range(1, totalRows+1):
    for c in range(1, totalColumns+1):
        print(sheet.cell(r,c).value, end='   ')
    print()

# Write data
workbook2 = openpyxl.load_workbook(path)
sheet2 = workbook2.active

for r in range(1,4):
    for c in range(1,3):
        sheet2.cell(r,c,'Welcome')
workbook2.save(path)


